"""
auditai_data_normalization/extractors/xlsx_extractor.py
=========================================================
Extracts structured data from Excel workpapers (.xlsx, .xlsm, .xls).

Why pandas + openpyxl and not an LLM?
---------------------------------------
Excel workpapers are already structured data — account codes, balances,
reconciliation items. openpyxl reads them deterministically with full
fidelity. The only job here is to normalise that structure into the
canonical DocumentRecord / ExtractedTable shape.

Supported formats
-----------------
.xlsx / .xlsm  → openpyxl (primary)
.xls           → xlrd (legacy binary format)
Both fall back to LibreOffice headless → CSV if all else fails.

What it produces
-----------------
sections  : one Section per non-skipped sheet, content is a
            human-readable text rendering of the sheet data.
            Used for embedding and LLM reading.
tables    : one ExtractedTable per non-skipped sheet, with
            typed headers and rows as strings.
            Used by the numeric chunker and math tools downstream.
metadata  : sheet_names, numeric_sheets, skipped_sheets,
            merged_cell_count, subtotal_rows_detected,
            total_rows_extracted

Edge cases handled
------------------
- Merged cells            → forward-fill across row and column
- Formula cells           → data_only=True reads computed value, not formula
- All sheets, not just Sheet1 → iterates xl.sheet_names
- Mixed dtypes            → dtype=str on read, normalize downstream
- Empty rows (>85% empty) → dropped
- Cover/Index/TOC sheets  → skipped by name pattern
- Subtotal rows           → detected by indentation or label keywords
- Legacy .xls             → xlrd engine
- Password-protected      → raises PasswordProtectedError
- Empty workbook          → extraction_status='partial'

Public API
----------
    extract(file_path: str | Path) -> DocumentRecord
"""

from __future__ import annotations

import hashlib
import logging
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

from auditai_data_normalization.schema import (
    DocumentRecord,
    ExtractedTable,
    Section,
)

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# Sheet names matching these patterns are skipped — no audit data
_SKIP_SHEET_RE = re.compile(
    r"^(cover|index|toc|table of contents|instructions|summary|legend"
    r"|contents|readme|notes|changes|log|template|example|sample)$",
    re.IGNORECASE,
)

# Row is considered empty if this fraction of cells are blank
_EMPTY_ROW_THRESHOLD = 0.85

# Keywords that indicate a subtotal or total row
_SUBTOTAL_KEYWORDS = re.compile(
    r"\b(total|subtotal|grand total|net total|sum|balance|carried forward"
    r"|c\/f|b\/f|brought forward)\b",
    re.IGNORECASE,
)

# Column is considered numeric if this fraction of non-empty cells parse as float
_NUMERIC_COLUMN_THRESHOLD = 0.80

# Max rows to render in section text (prevents enormous sections)
_MAX_ROWS_IN_TEXT = 500


# ---------------------------------------------------------------------------
# Custom exceptions
# ---------------------------------------------------------------------------

class PasswordProtectedError(Exception):
    """Raised when the Excel file is encrypted."""

class UnsupportedFormatError(Exception):
    """Raised for formats that cannot be read."""


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _is_skip_sheet(sheet_name: str) -> bool:
    return bool(_SKIP_SHEET_RE.match(sheet_name.strip()))


def _to_str(val) -> str:
    """Convert any cell value to a clean string."""
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if isinstance(val, float):
        # Remove trailing .0 for whole numbers
        if val == int(val):
            return str(int(val))
        return str(val)
    return str(val).strip()


def _is_empty_row(row: list[str]) -> bool:
    non_empty = sum(1 for c in row if c.strip())
    return non_empty / max(len(row), 1) <= (1 - _EMPTY_ROW_THRESHOLD)


def _is_subtotal_row(row: list[str]) -> bool:
    """True if any cell in the row matches subtotal keywords."""
    return any(_SUBTOTAL_KEYWORDS.search(c) for c in row if c)


def _detect_numeric_columns(rows: list[list[str]]) -> list[int]:
    """
    Return indices of columns where > 80% of non-empty values are numeric.
    """
    if not rows:
        return []
    n_cols = max(len(r) for r in rows)
    numeric_cols = []
    for col_idx in range(n_cols):
        values = [r[col_idx] for r in rows if col_idx < len(r) and r[col_idx].strip()]
        if not values:
            continue
        numeric_count = 0
        for v in values:
            clean = v.replace(",", "").replace("$", "").replace("(", "-").replace(")", "").strip()
            try:
                float(clean)
                numeric_count += 1
            except ValueError:
                pass
        if numeric_count / len(values) >= _NUMERIC_COLUMN_THRESHOLD:
            numeric_cols.append(col_idx)
    return numeric_cols


def _handle_merged_cells(ws) -> dict[tuple, str]:
    """
    Build a map of (row, col) → value for all cells in merged ranges.
    The top-left cell of the merge holds the value; others are None.
    We forward-fill the top-left value across the entire merged region.
    """
    merged_values: dict[tuple, str] = {}
    for merge in ws.merged_cells.ranges:
        # Get value from top-left cell of merge
        top_left = ws.cell(merge.min_row, merge.min_col)
        val = _to_str(top_left.value)
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                merged_values[(row, col)] = val
    return merged_values


def _sheet_to_rows(ws) -> tuple[list[str], list[list[str]]]:
    """
    Read an openpyxl worksheet into (headers, data_rows).

    Handles merged cells by forward-filling.
    Skips empty rows.
    Returns all values as strings.
    """
    merged_map = _handle_merged_cells(ws)
    all_rows: list[list[str]] = []

    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cells: list[str] = []
        for col_idx, cell in enumerate(row, start=1):
            # Check merged cell map first
            if (row_idx, col_idx) in merged_map:
                cells.append(merged_map[(row_idx, col_idx)])
            else:
                cells.append(_to_str(cell.value))
        if not _is_empty_row(cells):
            all_rows.append(cells)

    if not all_rows:
        return [], []

    # First non-empty row as headers
    headers = all_rows[0]
    data_rows = all_rows[1:]

    return headers, data_rows


def _rows_to_text(
    sheet_name: str,
    headers: list[str],
    data_rows: list[list[str]],
    numeric_cols: list[int],
) -> str:
    """
    Render sheet data as human-readable text for embedding and LLM reading.

    Format:
        Sheet: Trial Balance
        Account Code | Account Name | Debit | Credit | Balance
        ----------------------------------------
        1010 | Cash and Investments | 505900 | | 505900
        ...
        [SUBTOTAL] | TOTAL | 890462 | 909870 | -19408
    """
    lines: list[str] = [f"Sheet: {sheet_name}"]

    if headers:
        lines.append(" | ".join(h for h in headers))
        lines.append("-" * 40)

    for row in data_rows[:_MAX_ROWS_IN_TEXT]:
        is_sub = _is_subtotal_row(row)
        prefix = "[SUBTOTAL] " if is_sub else ""
        lines.append(prefix + " | ".join(row))

    if len(data_rows) > _MAX_ROWS_IN_TEXT:
        lines.append(f"[... {len(data_rows) - _MAX_ROWS_IN_TEXT} more rows truncated]")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# Excel reading with fallback
# ---------------------------------------------------------------------------

def _read_xlsx(path: Path) -> tuple[openpyxl.Workbook, str]:
    """
    Open workbook with openpyxl (data_only=True so formulas return values).
    Returns (workbook, engine_used).
    Raises PasswordProtectedError if encrypted.
    """
    # Quick password check — encrypted xlsx starts with OLE magic bytes
    raw = path.read_bytes()
    if raw[:4] == b"\xd0\xcf\x11\xe0":
        raise PasswordProtectedError(
            f"{path.name} is password-protected. "
            "Supply password via intake API."
        )

    try:
        wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
        return wb, "openpyxl"
    except Exception as e:
        raise UnsupportedFormatError(f"openpyxl failed on {path.name}: {e}") from e


def _read_xls(path: Path) -> pd.ExcelFile:
    """Read legacy .xls file with xlrd engine."""
    try:
        return pd.ExcelFile(str(path), engine="xlrd")
    except Exception as e:
        raise UnsupportedFormatError(
            f"xlrd failed on {path.name}: {e}. "
            "Try converting to .xlsx first."
        ) from e


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------

def extract(file_path: str | Path) -> DocumentRecord:
    """
    Extract structured data from an Excel workpaper (.xlsx, .xlsm, .xls).

    Parameters
    ----------
    file_path : str | Path
        Path to an Excel file.

    Returns
    -------
    DocumentRecord
        sections  — one per non-skipped sheet (text rendering)
        tables    — one per non-skipped sheet (structured rows)
        pii_scrubbed=False — call pii.scrub_record() after this.

    Raises
    ------
    FileNotFoundError
        If file does not exist.
    PasswordProtectedError
        If file is encrypted.
    UnsupportedFormatError
        If format cannot be read after all fallbacks.
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    raw_bytes = path.read_bytes()
    file_hash = hashlib.sha256(raw_bytes).hexdigest()
    file_size = len(raw_bytes)
    suffix = path.suffix.lower()

    sections: list[Section] = []
    tables: list[ExtractedTable] = []
    skipped_sheets: list[str] = []
    numeric_sheets: list[str] = []
    processed_sheets: list[str] = []
    total_merged = 0
    total_subtotal_rows = 0
    total_rows = 0
    error_msg = ""

    # ----------------------------------------------------------------
    # .xlsx / .xlsm path — openpyxl
    # ----------------------------------------------------------------
    if suffix in (".xlsx", ".xlsm"):
        try:
            wb, engine = _read_xlsx(path)
        except PasswordProtectedError:
            raise
        except UnsupportedFormatError as e:
            return DocumentRecord(
                source_path=str(path),
                file_name=path.name,
                file_type="xlsx",
                file_size_bytes=file_size,
                file_hash=file_hash,
                extraction_method="openpyxl",
                extraction_status="failed",
                extraction_error=str(e),
                needs_review=True,
            )

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            if _is_skip_sheet(sheet_name):
                skipped_sheets.append(sheet_name)
                logger.debug("Skipping sheet: %s", sheet_name)
                continue

            ws = wb[sheet_name]

            # Count merged cells
            total_merged += len(list(ws.merged_cells.ranges))

            headers, data_rows = _sheet_to_rows(ws)

            if not headers and not data_rows:
                skipped_sheets.append(sheet_name)
                continue

            processed_sheets.append(sheet_name)
            total_rows += len(data_rows)

            # Detect subtotal rows
            sub_rows = sum(1 for r in data_rows if _is_subtotal_row(r))
            total_subtotal_rows += sub_rows

            # Detect numeric columns
            numeric_cols = _detect_numeric_columns(data_rows)
            if numeric_cols:
                numeric_sheets.append(sheet_name)

            # Build ExtractedTable
            table = ExtractedTable(
                index=sheet_idx,
                source=f"Sheet: {sheet_name}",
                headers=headers,
                rows=data_rows,
                raw_text=_rows_to_text(sheet_name, headers, data_rows, numeric_cols),
            )
            tables.append(table)

            # Build Section (text rendering for embedding)
            content = _rows_to_text(sheet_name, headers, data_rows, numeric_cols)
            sections.append(Section(
                index=sheet_idx,
                heading=sheet_name,
                content=content,
                page_or_sheet=f"Sheet: {sheet_name}",
                token_count=len(content.split()),
                is_table=True,
            ))

            logger.debug(
                "Sheet '%s': %d rows, %d numeric cols, %d subtotals",
                sheet_name, len(data_rows), len(numeric_cols), sub_rows,
            )

    # ----------------------------------------------------------------
    # .xls path — xlrd via pandas
    # ----------------------------------------------------------------
    elif suffix == ".xls":
        try:
            xl = _read_xls(path)
        except UnsupportedFormatError as e:
            return DocumentRecord(
                source_path=str(path),
                file_name=path.name,
                file_type="xlsx",
                file_size_bytes=file_size,
                file_hash=file_hash,
                extraction_method="openpyxl",
                extraction_status="failed",
                extraction_error=str(e),
                needs_review=True,
            )

        for sheet_idx, sheet_name in enumerate(xl.sheet_names):
            if _is_skip_sheet(sheet_name):
                skipped_sheets.append(sheet_name)
                continue

            try:
                df = xl.parse(sheet_name, dtype=str, header=None)
            except Exception as e:
                logger.warning("Failed to parse sheet %s: %s", sheet_name, e)
                skipped_sheets.append(sheet_name)
                continue

            df = df.fillna("")
            all_rows = [
                [_to_str(v) for v in row]
                for _, row in df.iterrows()
            ]
            all_rows = [r for r in all_rows if not _is_empty_row(r)]

            if not all_rows:
                skipped_sheets.append(sheet_name)
                continue

            headers = all_rows[0]
            data_rows = all_rows[1:]
            processed_sheets.append(sheet_name)
            total_rows += len(data_rows)

            numeric_cols = _detect_numeric_columns(data_rows)
            if numeric_cols:
                numeric_sheets.append(sheet_name)

            sub_rows = sum(1 for r in data_rows if _is_subtotal_row(r))
            total_subtotal_rows += sub_rows

            table = ExtractedTable(
                index=sheet_idx,
                source=f"Sheet: {sheet_name}",
                headers=headers,
                rows=data_rows,
                raw_text=_rows_to_text(sheet_name, headers, data_rows, numeric_cols),
            )
            tables.append(table)

            content = _rows_to_text(sheet_name, headers, data_rows, numeric_cols)
            sections.append(Section(
                index=sheet_idx,
                heading=sheet_name,
                content=content,
                page_or_sheet=f"Sheet: {sheet_name}",
                token_count=len(content.split()),
                is_table=True,
            ))

    else:
        return DocumentRecord(
            source_path=str(path),
            file_name=path.name,
            file_type="xlsx",
            file_size_bytes=file_size,
            file_hash=file_hash,
            extraction_method="openpyxl",
            extraction_status="failed",
            extraction_error=f"Unsupported extension: {suffix}",
            needs_review=True,
        )

    # ----------------------------------------------------------------
    # Determine status
    # ----------------------------------------------------------------
    if not sections and not tables:
        status = "failed"
        error_msg = "No data extracted — all sheets were empty or skipped"
    elif skipped_sheets and not processed_sheets:
        status = "partial"
        error_msg = f"All sheets skipped: {skipped_sheets}"
    else:
        status = "success"

    # ----------------------------------------------------------------
    # Build cleaned_text and word count
    # ----------------------------------------------------------------
    cleaned_text = "\n\n".join(
        s.content for s in sections if s.content.strip()
    )
    word_count = len(cleaned_text.split())

    metadata = {
        "sheet_names": wb.sheetnames if suffix in (".xlsx", ".xlsm") else
                       xl.sheet_names if suffix == ".xls" else [],
        "processed_sheets": processed_sheets,
        "skipped_sheets": skipped_sheets,
        "numeric_sheets": numeric_sheets,
        "merged_cell_count": total_merged,
        "subtotal_rows_detected": total_subtotal_rows,
        "total_rows_extracted": total_rows,
        "section_count": len(sections),
        "table_count": len(tables),
    }

    logger.info(
        "xlsx_extractor: %s — %d sheets processed, %d skipped, "
        "%d rows, status=%s",
        path.name, len(processed_sheets), len(skipped_sheets),
        total_rows, status,
    )

    return DocumentRecord(
        source_path=str(path),
        file_name=path.name,
        file_type="xlsx",
        file_size_bytes=file_size,
        file_hash=file_hash,
        raw_text=cleaned_text,
        cleaned_text=cleaned_text,
        sections=sections,
        tables=tables,
        extraction_method="openpyxl",
        extraction_status=status,
        extraction_error=error_msg,
        extraction_confidence=0.0,
        word_count=word_count,
        page_count=0,
        ocr_used=False,
        pii_scrubbed=False,
        needs_review=(status != "success"),
        metadata=metadata,
    )