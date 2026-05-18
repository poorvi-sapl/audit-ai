"""
raw_to_training_pair/extraction/xlsx_extractor.py
===================================================
Phase 2 raw extractor for .xlsx, .xlsm, and .xls files.

Contract
--------
Returns a plain dict — no DocumentRecord, no normalization.

Output shape
------------
{
    "text"    : str,   # all sheets rendered as readable text, joined
    "tables"  : list,  # [{"sheet": str, "headers": [...], "rows": [[...]]}, ...]
    "metadata": dict   # sheet_names, processed_sheets, skipped_sheets, row_count
}

Public API
----------
    extract(file_path: str | Path) -> dict
"""

from __future__ import annotations

import logging
import re
from pathlib import Path

import numpy as np
import openpyxl
import pandas as pd

logger = logging.getLogger(__name__)

# Sheet names matching these are skipped
_SKIP_SHEET_RE = re.compile(
    r"^(cover|index|toc|table of contents|instructions|summary|legend"
    r"|contents|readme|notes|changes|log|template|example|sample)$",
    re.IGNORECASE,
)

_EMPTY_ROW_THRESHOLD = 0.85
_MAX_ROWS_IN_TEXT = 500


def _is_skip_sheet(name: str) -> bool:
    return bool(_SKIP_SHEET_RE.match(name.strip()))


def _to_str(val) -> str:
    if val is None or (isinstance(val, float) and np.isnan(val)):
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _is_empty_row(row: list[str]) -> bool:
    non_empty = sum(1 for c in row if c.strip())
    return non_empty / max(len(row), 1) <= (1 - _EMPTY_ROW_THRESHOLD)


def _sheet_to_rows(ws) -> tuple[list[str], list[list[str]]]:
    """Read openpyxl worksheet into (headers, data_rows)."""
    # Handle merged cells
    merged_map: dict[tuple, str] = {}
    for merge in ws.merged_cells.ranges:
        top_left = ws.cell(merge.min_row, merge.min_col)
        val = _to_str(top_left.value)
        for row in range(merge.min_row, merge.max_row + 1):
            for col in range(merge.min_col, merge.max_col + 1):
                merged_map[(row, col)] = val

    all_rows = []
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        cells = []
        for col_idx, cell in enumerate(row, start=1):
            if (row_idx, col_idx) in merged_map:
                cells.append(merged_map[(row_idx, col_idx)])
            else:
                cells.append(_to_str(cell.value))
        if not _is_empty_row(cells):
            all_rows.append(cells)

    if not all_rows:
        return [], []

    return all_rows[0], all_rows[1:]


def _rows_to_text(sheet_name: str, headers: list[str], rows: list[list[str]]) -> str:
    lines = [f"Sheet: {sheet_name}", " | ".join(headers), "-" * 40]
    for row in rows[:_MAX_ROWS_IN_TEXT]:
        lines.append(" | ".join(row))
    if len(rows) > _MAX_ROWS_IN_TEXT:
        lines.append(f"[... {len(rows) - _MAX_ROWS_IN_TEXT} more rows truncated]")
    return "\n".join(lines)


def extract(file_path: str | Path) -> dict:
    """
    Extract raw rows and text from an Excel file.

    Parameters
    ----------
    file_path : str | Path

    Returns
    -------
    dict with keys: text, tables, metadata

    Raises
    ------
    FileNotFoundError
        If file does not exist.
    RuntimeError
        If openpyxl/xlrd cannot open the file.
    """
    path = Path(file_path)

    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")

    suffix = path.suffix.lower()
    tables = []
    text_blocks = []
    processed_sheets = []
    skipped_sheets = []

    # --- .xlsx / .xlsm via openpyxl ---
    if suffix in (".xlsx", ".xlsm"):
        try:
            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=False)
        except Exception as e:
            raise RuntimeError(f"openpyxl failed on {path.name}: {e}") from e

        for sheet_name in wb.sheetnames:
            if _is_skip_sheet(sheet_name):
                skipped_sheets.append(sheet_name)
                continue
            ws = wb[sheet_name]
            headers, rows = _sheet_to_rows(ws)
            if not headers and not rows:
                skipped_sheets.append(sheet_name)
                continue
            processed_sheets.append(sheet_name)
            tables.append({"sheet": sheet_name, "headers": headers, "rows": rows})
            text_blocks.append(_rows_to_text(sheet_name, headers, rows))

        sheet_names = wb.sheetnames

    # --- .xls via xlrd/pandas ---
    elif suffix == ".xls":
        try:
            xl = pd.ExcelFile(str(path), engine="xlrd")
        except Exception as e:
            raise RuntimeError(f"xlrd failed on {path.name}: {e}") from e

        for sheet_name in xl.sheet_names:
            if _is_skip_sheet(sheet_name):
                skipped_sheets.append(sheet_name)
                continue
            try:
                df = xl.parse(sheet_name, dtype=str, header=None).fillna("")
            except Exception:
                skipped_sheets.append(sheet_name)
                continue

            all_rows = [
                [_to_str(v) for v in row]
                for _, row in df.iterrows()
                if not _is_empty_row([_to_str(v) for v in row])
            ]
            if not all_rows:
                skipped_sheets.append(sheet_name)
                continue

            headers, rows = all_rows[0], all_rows[1:]
            processed_sheets.append(sheet_name)
            tables.append({"sheet": sheet_name, "headers": headers, "rows": rows})
            text_blocks.append(_rows_to_text(sheet_name, headers, rows))

        sheet_names = xl.sheet_names

    else:
        raise RuntimeError(f"Unsupported extension: {suffix}")

    full_text = "\n\n".join(text_blocks)
    total_rows = sum(len(t["rows"]) for t in tables)

    logger.info(
        "p2/xlsx_extractor: %s — %d sheets, %d rows",
        path.name, len(processed_sheets), total_rows,
    )

    return {
        "text": full_text,
        "tables": tables,
        "metadata": {
            "sheet_names": list(sheet_names),
            "processed_sheets": processed_sheets,
            "skipped_sheets": skipped_sheets,
            "row_count": total_rows,
        },
    }